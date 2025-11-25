from django.shortcuts import render, redirect
from django.db.models import Count, Sum, Q
from django.views.decorators.csrf import csrf_exempt
from django.contrib import messages
from django.http import JsonResponse
from django.template.loader import render_to_string
import subprocess
import os
import fitz
from ..models import Lectura, RegistroUsuarios, EvaluacionLecturaIndividual, EvaluacionLectura, RegistroAdmin, VistaAdmin

VAL_MAX = 3
def porcentaje_final(e):
    """
    Calcula el porcentaje ponderado de UNA lectura individual replicando
    la lógica de guardar_respuesta:
      - porcentaje_inferencia = (puntaje / VAL_MAX) * 100
      - puntaje_velocidad: 50 | 75 | 100 según palabras_por_minuto
      - resultado = porcentaje_inferencia*0.80 + puntaje_velocidad*0.20
    """
    if e.puntaje is None:
        return 0.0
    try:
        porcentaje_inferencia = (float(e.puntaje) / float(VAL_MAX)) * 100.0
    except Exception:
        porcentaje_inferencia = 0.0

    ppm = e.palabras_por_minuto or 0
    puntaje_velocidad = 50
    if ppm >= 230:
        puntaje_velocidad = 100
    elif ppm >= 150:
        puntaje_velocidad = 75

    return (porcentaje_inferencia * 0.80) + (puntaje_velocidad * 0.20)

# Vista: Dashboard del administrador
def dashboard_admin(request):
    usuarios_activos = RegistroUsuarios.objects.annotate(
        evaluaciones_realizadas=Count(
            'evaluacionlecturaindividual',
            filter=Q(evaluacionlecturaindividual__puntaje__isnull=False)
        )
    ).filter(evaluaciones_realizadas__gt=0)

    total_usuarios = usuarios_activos.count()

    tipos = ["Argumentativo", "Descriptivo", "Expositivo", "Narrativo"]
    resumen = []

    for tipo in tipos:
        lecturas = EvaluacionLecturaIndividual.objects.filter(tipo_texto=tipo)
        total_lecturas = lecturas.count()
        porcentajes = [porcentaje_final(e) for e in lecturas if e.puntaje is not None]
        promedio_pct = (sum(porcentajes) / len(porcentajes)) if porcentajes else 0.0

        resumen.append({
            "tipo": tipo,
            "total_lecturas": total_lecturas,
            "promedio_porcentaje": round(promedio_pct, 2)
        })

    contexto = {
        "total_usuarios": total_usuarios,
        "resumen_por_tipo": resumen,
    }
    return render(request, "evaluacionescl/dashboard_admin.html", contexto)


# Vista: Resultados globales por tipo de texto
def admin_resultados(request):
    tipos = ["Argumentativo", "Descriptivo", "Expositivo", "Narrativo"]
    resumen = []

    for tipo in tipos:
        lecturas = EvaluacionLecturaIndividual.objects.filter(tipo_texto=tipo)

        # Nuevo: usa porcentaje_final para cada lectura
        porcentajes = [porcentaje_final(e) for e in lecturas if e.puntaje is not None]
        cantidad = len(porcentajes)
        promedio_pct = (sum(porcentajes) / cantidad) if cantidad > 0 else 0.0

        if cantidad > 0:
            pct = promedio_pct
            if pct >= 90:
                nivel = f"Alto (Comprensión profunda) - {int(pct)}%"
            elif pct >= 60:
                nivel = f"Medio (Comprensión adecuada) - {int(pct)}%"
            elif pct >= 30:
                nivel = f"Bajo (Comprensión superficial) - {int(pct)}%"
            else:
                nivel = f"Deficiente (No comprensión) - {int(pct)}%"
        else:
            nivel = "Sin evaluar"

        resumen.append({
            "tipo": tipo,
            "cantidad": cantidad,
            # opcional: si tu template lo muestra, puedes mantener un “puntaje_total” estimado
            # convirtiendo desde porcentaje a escala 0–VAL_MAX:
            "puntaje_total": round((sum(porcentajes) / 100.0) * VAL_MAX, 2) if cantidad else 0,
            "promedio_porcentaje": round(promedio_pct, 2),
            "nivel": nivel,
        })

    total_usuarios = RegistroUsuarios.objects.count()

    return render(request, 'evaluacionescl/admin_resultados.html', {
        'resumen': resumen,
        'total_usuarios': total_usuarios
    })

# Vista: Estadísticas por alumno con búsqueda
from django.db.models import Q
from django.shortcuts import render, get_object_or_404
from ..models import RegistroUsuarios, EvaluacionLecturaIndividual
from .evaluacion_views import calcular_porcentaje

VAL_MAX = 3

def admin_estadisticas(request):
    query = request.GET.get('q', '')
    usuarios = RegistroUsuarios.objects.all()
    if query:
        usuarios = usuarios.filter(
            Q(nombre__icontains=query) |
            Q(apellido__icontains=query) |
            Q(matricula__icontains=query)
        )

    resultados = []
    for user in usuarios:
        lecturas = EvaluacionLecturaIndividual.objects.filter(usuario=user)
        # Usar porcentaje_final por lectura (ya ponderado); ignorar las que no tienen puntaje
        porcentajes = [porcentaje_final(e) for e in lecturas if e.puntaje is not None]
        total_textos = len(porcentajes)

        promedio_pct = (sum(porcentajes) / total_textos) if total_textos else 0.0

        if total_textos > 0:
            pct = promedio_pct
            if pct >= 90:
                nivel = f"Alto (Comprensión profunda) - {int(round(pct))}%"
            elif pct >= 60:
                nivel = f"Medio (Comprensión adecuada) - {int(round(pct))}%"
            elif pct >= 30:
                nivel = f"Bajo (Comprensión superficial) - {int(round(pct))}%"
            else:
                nivel = f"Deficiente (No comprensión) - {int(round(pct))}%"
        else:
            nivel = "Sin evaluar"

        # Conserva “puntaje” si tu template lo muestra, derivado desde % (escala 0–VAL_MAX)
        puntaje_equivalente = round((sum(porcentajes) / 100.0) * VAL_MAX, 2) if total_textos else 0

        resultados.append({
            "id": user.id,
            "nombre": f"{user.nombre} {user.apellido}",
            "matricula": user.matricula,
            "textos": total_textos,
            "puntaje": puntaje_equivalente,             # si la plantilla usa “puntaje”
            "porcentaje": round(promedio_pct, 2),       # úsalo si la plantilla muestra porcentaje
            "nivel": nivel
        })

    return render(request, 'evaluacionescl/admin_estadisticas.html', {
        'resultados': resultados,
        'query': query
    })

#--------------------------------------------------------------------------------------------

from django.contrib import messages
from django.shortcuts import redirect

def eliminar_usuario(request, usuario_id):
    if request.method == "POST":
        try:
            usuario = RegistroUsuarios.objects.get(id=usuario_id)
            usuario.delete()
            messages.success(request, "Usuario eliminado correctamente.")
        except RegistroUsuarios.DoesNotExist:
            messages.error(request, "El usuario no existe.")
    return redirect("admin_estadisticas")

#----------------------------------------------------------------------------------------------

from django.shortcuts import get_object_or_404, render, redirect
from ..models import RegistroUsuarios, EvaluacionLecturaIndividual
from .evaluacion_views import calcular_porcentaje  # Asegúrate de importar si está separado


def ver_resultados_alumno(request, usuario_id):
    alumno = get_object_or_404(RegistroUsuarios, id=usuario_id)
    tipos = ["Argumentativo", "Descriptivo", "Expositivo", "Narrativo"]
    resultados = []

    for tipo in tipos:
        lecturas = EvaluacionLecturaIndividual.objects.filter(usuario=alumno, tipo_texto=tipo)
        # Usa porcentaje_final por lectura (no promedies puntaje crudo)
        porcentajes = [porcentaje_final(e) for e in lecturas if e.puntaje is not None]
        total = len(porcentajes)

        if total:
            promedio_pct = sum(porcentajes) / total
            porcentaje = promedio_pct
            if porcentaje >= 90:
                nivel = f"Alto (Comprensión profunda) - {int(round(porcentaje))}%"
            elif porcentaje >= 60:
                nivel = f"Medio (Comprensión adecuada) - {int(round(porcentaje))}%"
            elif porcentaje >= 30:
                nivel = f"Bajo (Comprensión superficial) - {int(round(porcentaje))}%"
            else:
                nivel = f"Deficiente (No comprensión) - {int(round(porcentaje))}%"
        else:
            porcentaje = 0
            nivel = "Sin evaluar"

        resultados.append({
            "tipo": tipo,
            "cantidad": total,
            "nivel": nivel,
            "porcentaje": round(porcentaje, 2)
        })

    # … deja igual la lógica de la gráfica global mensual …
    from collections import defaultdict
    from django.utils.timezone import localtime
    from datetime import datetime

    individuales = EvaluacionLecturaIndividual.objects.filter(
        usuario=alumno,
        respuesta_usuario__isnull=False,
        puntaje__isnull=False
    )

    mes_tipo_set = defaultdict(set)
    for indiv in individuales:
        fecha_local = indiv.fecha_lectura
        mes_clave = fecha_local.strftime("%Y-%m")
        mes_tipo_set[mes_clave].add(indiv.tipo_texto)

    resumenes = EvaluacionLectura.objects.filter(usuario=alumno)
    porcentajes_por_tipo = {r.tipo_texto: r.porcentaje or 0 for r in resumenes}

    meses = []
    promedios_globales = []
    meses_esp = {
        "Jan": "Ene", "Feb": "Feb", "Mar": "Mar", "Apr": "Abr", "May": "May",
        "Jun": "Jun", "Jul": "Jul", "Aug": "Ago", "Sep": "Sep",
        "Oct": "Oct", "Nov": "Nov", "Dec": "Dic"
    }

    for mes in sorted(mes_tipo_set.keys()):
        suma = sum([porcentajes_por_tipo.get(t, 0) for t in tipos])
        promedio = suma / 4
        promedios_globales.append(round(promedio, 1))

        fecha_obj = datetime.strptime(mes, "%Y-%m")
        mes_abbr = fecha_obj.strftime("%b")
        mes_nombre = meses_esp.get(mes_abbr, mes_abbr)
        meses.append(f"{mes_nombre} {fecha_obj.year}")

    if promedios_globales:
        ultimo = promedios_globales[-1]
        if ultimo >= 90:
            nivel_global = "Alto (Comprensión profunda)"
        elif ultimo >= 60:
            nivel_global = "Medio (Comprensión adecuada)"
        elif ultimo >= 30:
            nivel_global = "Bajo (Comprensión superficial)"
        else:
            nivel_global = "Deficiente (No comprensión)"
    else:
        nivel_global = "Sin evaluar"

    return render(request, "evaluacionescl/ver_resultados_alumno.html", {
        "resultados": resultados,
        "alumno": alumno,
        "usuario_id": alumno.id,
        "nombre_alumno": alumno.nombre,
        "meses": meses,
        "promedios": promedios_globales,
        "nivel_global": nivel_global
    })


from django.shortcuts import render, get_object_or_404
from ..models import RegistroUsuarios, EvaluacionLecturaIndividual

def calcular_porcentaje(promedio):
    if promedio is None:
        return 0
    return (promedio / VAL_MAX) * 100

# ✅ 4. ver_grafica_alumno_tipo (con título, fecha y hora en líneas separadas)
# ✅ 4. ver_grafica_alumno_tipo (con multilínea real para Chart.js)

def ver_grafica_alumno_tipo(request, usuario_id, tipo_texto):
    alumno = get_object_or_404(RegistroUsuarios, id=usuario_id)
    lecturas = EvaluacionLecturaIndividual.objects.filter(
        usuario=alumno,
        tipo_texto=tipo_texto
    ).order_by('fecha_lectura')

    if not lecturas.exists():
        return render(request, "evaluacionescl/no_evaluaciones_admin.html", {
            "mensaje": f"No hay lecturas evaluadas para el tipo '{tipo_texto}' de este alumno."
        })

    titulos, porcentajes, tipos_inferencia, tooltips = [], [], [], []
    vistos = set()

    for l in lecturas:
        if l.titulo_lectura not in vistos:
            vistos.add(l.titulo_lectura)

            from django.utils.timezone import localtime
            fecha_local = l.fecha_lectura
            meses = {
                "Jan": "Ene", "Feb": "Feb", "Mar": "Mar", "Apr": "Abr",
                "May": "May", "Jun": "Jun", "Jul": "Jul", "Aug": "Ago",
                "Sep": "Sep", "Oct": "Oct", "Nov": "Nov", "Dec": "Dic"
            }
            mes_abbr = fecha_local.strftime("%b")
            mes_esp = meses.get(mes_abbr, mes_abbr)
            fecha_linea1 = f"{fecha_local.strftime('%d')}/{mes_esp}/{fecha_local.strftime('%Y')}"
            fecha_linea2 = fecha_local.strftime("%H:%M")
            titulo = l.titulo_lectura
            if len(titulo) > 20:
                titulo = titulo[:20] + "…"
            etiqueta = f"{titulo}\n{fecha_linea1}\n{fecha_linea2}"
            titulos.append(etiqueta)

            porcentaje = porcentaje_final(l) if l.puntaje is not None else 0
            porcentajes.append(porcentaje)
            tooltips.append(f"{int(porcentaje)}%" if l.puntaje is not None else "Sin evaluar")

            tipo = l.tipo_inferencia
            if tipo == "no_inferencia_sinsentido":
                tipo = "No inferencia: sin sentido"
            elif tipo == "no_inferencia_parafrasis":
                tipo = "No inferencia: paráfrasis"
            elif tipo:
                tipo = tipo.capitalize()
            else:
                tipo = "No inferencia: sin sentido"

            tipos_inferencia.append(tipo)

    return render(request, "evaluacionescl/ver_grafica_alumno_tipo.html", {
        "tipo": tipo_texto,
        "titulos": titulos,
        "porcentajes": porcentajes,
        "tipos_inferencia": tipos_inferencia,
        "tooltips": tooltips,
        "alumno_id": alumno.id  # 👈 ESTA LÍNEA SOLUCIONA TU ERROR
    })







# Exportar estadisticas admin
from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter
from ..models import EvaluacionLecturaIndividual

def exportar_admin_estadisticas_excel(request):
    from ..models import RegistroUsuarios
    from .evaluacion_views import calcular_porcentaje

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Estadísticas por alumno"

    headers = ["Nombre", "Matrícula", "Textos leídos", "Puntaje total", "Porcentaje", "Nivel de comprensión"]
    ws.append(headers)

    usuarios = RegistroUsuarios.objects.all()
    for user in usuarios:
        lecturas = EvaluacionLecturaIndividual.objects.filter(usuario=user)
        total = lecturas.count()
        porcentajes = [porcentaje_final(e) for e in lecturas if e.puntaje is not None]
        promedio_pct = (sum(porcentajes) / len(porcentajes)) if porcentajes else 0.0
        porcentaje = promedio_pct
        # Si quieres conservar “puntaje_total”, puedes derivarlo del %:
        puntaje_total = round((sum(porcentajes) / 100.0) * VAL_MAX, 2) if total else 0

        if total:
            if porcentaje >= 90:
                nivel = "Alto (Comprensión profunda)"
            elif porcentaje >= 60:
                nivel = "Medio (Comprensión adecuada)"
            elif porcentaje >= 30:
                nivel = "Bajo (Comprensión superficial)"
            else:
                nivel = "Deficiente (No comprensión)"
        else:
            nivel = "Sin evaluar"

        fila = [
            f"{user.nombre} {user.apellido}",
            user.matricula,
            total,
            puntaje_total,
            int(porcentaje),
            nivel
        ]
        ws.append(fila)

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].auto_size = True

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename=admin_estadisticas_alumnos.xlsx'
    wb.save(response)
    return response


# Vista para exportar resultados globales
from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter
from ..models import EvaluacionLecturaIndividual

VAL_MAX = 3  # Si no está definido arriba, agrégalo

def exportar_admin_resultados_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumen por tipo de texto"

    headers = ["Tipo de texto", "Documentos leídos", "Puntaje total", "Porcentaje", "Nivel de comprensión"]
    ws.append(headers)

    tipos = ["Argumentativo", "Descriptivo", "Expositivo", "Narrativo"]

    for tipo in tipos:
        lecturas = EvaluacionLecturaIndividual.objects.filter(tipo_texto=tipo)
        porcentajes = [porcentaje_final(e) for e in lecturas if e.puntaje is not None]
        cantidad = len(porcentajes)
        promedio_pct = (sum(porcentajes) / cantidad) if cantidad else 0.0
        porcentaje = promedio_pct
        suma = round((sum(porcentajes) / 100.0) * VAL_MAX, 2) if cantidad else 0

        if cantidad > 0:
            if porcentaje >= 90:
                nivel = f"Alto (Comprensión profunda) - {int(porcentaje)}%"
            elif porcentaje >= 60:
                nivel = f"Medio (Comprensión adecuada) - {int(porcentaje)}%"
            elif porcentaje >= 30:
                nivel = f"Bajo (Comprensión superficial) - {int(porcentaje)}%"
            else:
                nivel = f"Deficiente (No comprensión) - {int(porcentaje)}%"
        else:
            nivel = "Sin evaluar"

        fila = [tipo, cantidad, suma, int(porcentaje), nivel]
        ws.append(fila)

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].auto_size = True

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename=admin_resultados_globales.xlsx'
    wb.save(response)
    return response



#-----------------------------------------------------------------------------
from django.db import connection
from django.shortcuts import render, redirect
from django.contrib import messages
from ..models import RegistroAdmin, RegistroUsuarios, EvaluacionLectura, EvaluacionLecturaIndividual, VistaAdmin

def reset_auto_increment(tabla):
    with connection.cursor() as cursor:
        cursor.execute(f"ALTER TABLE {tabla} AUTO_INCREMENT = 1")

def resetear_datos(request):
    if 'admin_id' not in request.session:
        return redirect('login_admin')  # validación personalizada
    if request.method == "POST":
        if "reset_todo" in request.POST:
            # Eliminar todo, incluyendo el admin actual
            RegistroUsuarios.objects.all().delete()
            RegistroAdmin.objects.all().delete()
            EvaluacionLectura.objects.all().delete()
            EvaluacionLecturaIndividual.objects.all().delete()
            VistaAdmin.objects.all().delete()

            reset_auto_increment("evaluacionescl_registrousuarios")
            reset_auto_increment("evaluacionescl_registroadmin")
            reset_auto_increment("evaluacionescl_evaluacionlectura")
            reset_auto_increment("evaluacionescl_evaluacionlecturaindividual")
            reset_auto_increment("evaluacionescl_vistaadmin")

            # Guardar mensaje antes de limpiar sesión
            add_message(request, messages.SUCCESS, "⚠️ Se reseteó TODO, incluyendo el admin en sesión.")

            # Limpiar la sesión y redirigir
            request.session.flush()
            return redirect("login_admin")

        elif "cancelar" in request.POST:
            messages.info(request, "🚫 Operación cancelada.")
            return redirect("dashboard_admin")

    return render(request, "evaluacionescl/resetear_datos_confirmacion.html")


#---------------------------------------------------------------------------------
# subir pdfs

import os
from django.conf import settings
from django.contrib import messages
from django.shortcuts import render, redirect

def subir_pdf(request):
    if request.method == 'POST':
        # Verificamos si se enviaron archivos en la petición
        if not request.FILES:
            messages.error(request, "No se seleccionó ningún archivo.")
            return redirect('subir_pdf')

        subidos_ok = 0
        # Iteramos sobre cada archivo que el usuario subió
        for input_name, archivo_subido in request.FILES.items():
            
            # Extraemos el tipo de texto (Argumentativo, etc.) del nombre del input
            try:
                tipo_texto = input_name.replace('archivo_', '').capitalize()
            except:
                continue # Ignoramos inputs con nombres inesperados

            # Creamos una instancia de nuestro nuevo modelo 'Lectura'
            nueva_lectura = Lectura(
                titulo=archivo_subido.name.replace('.pdf', '').replace('_', ' '),
                tipo_texto=tipo_texto,
                archivo_pdf=archivo_subido
            )
            
            try:
                # Guardamos el objeto en la base de datos.
                # Django se encarga de guardar el archivo físico en la carpeta correcta.
                nueva_lectura.save()
                
                # --- INICIA LA LÓGICA PARA CONTAR PALABRAS ---
                texto_completo = ""
                # Abrimos el PDF que Django acaba de guardar usando su ruta
                with fitz.open(nueva_lectura.archivo_pdf.path) as doc:
                    for pagina in doc:
                        texto_completo += pagina.get_text("text")
                
                conteo = len(texto_completo.split())
                nueva_lectura.conteo_palabras = conteo
                # Actualizamos el registro en la BD con el número de palabras
                nueva_lectura.save(update_fields=['conteo_palabras'])
                # --- FIN DE LA LÓGICA ---

                messages.success(request, f"✅ '{archivo_subido.name}' subido y procesado ({conteo} palabras).")
                subidos_ok += 1

            except Exception as e:
                messages.error(request, f"❌ Error al procesar '{archivo_subido.name}': {e}")
        
        return redirect('subir_pdf')

    # Para el método GET, simplemente mostramos la página de subida
    return render(request, 'evaluacionescl/subir_pdf.html')

